import streamlit as st
import pandas as pd
import os
import time
import matplotlib.pyplot as plt

LOG_FILE = "/home/localuser/dlp-llm-proxy/dlp-proxy/logs/mitmproxy_events.csv"

st.set_page_config(
    page_title="DLP Proxy Dashboard",
    layout="wide"
)

# -----------------------------
# Cyber Theme Styling
# -----------------------------
st.markdown("""
<style>
    /* -----------------------------
       GLOBAL THEME
    ----------------------------- */
    .stApp {
        background-color: #0f172a;
        color: #e2e8f0;
    }

    h1, h2, h3, h4 {
        color: #f8fafc;
    }

    /* -----------------------------
       SIDEBAR STYLING
    ----------------------------- */
    section[data-testid="stSidebar"] {
        background-color: #111827;
    }

    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] h4 {
        color: #ffffff !important;
    }

    section[data-testid="stSidebar"] label {
        color: #ffffff !important;
    }

    section[data-testid="stSidebar"] span {
        color: #ffffff !important;
    }

    section[data-testid="stSidebar"] * {
        color: #ffffff !important;
    }

    /* -----------------------------
       METRIC CARDS
    ----------------------------- */
    div[data-testid="metric-container"] {
        background-color: #0b5ed7;
        border: 1px solid #60a5fa;
        padding: 16px;
        border-radius: 14px;
    }

    [data-testid="stMetricLabel"] p {
        color: #ffffff !important;
    }

    [data-testid="stMetricValue"] {
        color: #ffffff !important;
    }

    /* -----------------------------
       TABS
    ----------------------------- */
    button[data-baseweb="tab"] {
        color: #ffffff !important;
    }

    button[aria-selected="true"] {
        color: #ffffff !important;
        border-bottom: 2px solid #ef4444 !important;
    }

    button[aria-selected="false"] {
        color: #93c5fd !important;
    }

    /* -----------------------------
       DATAFRAME
    ----------------------------- */
    div[data-testid="stDataFrame"] {
        border: 1px solid #60a5fa;
        border-radius: 12px;
        overflow: hidden;
    }

    /* -----------------------------
       SMALL TEXT
    ----------------------------- */
    .small-note {
        color: #67e8f9;
        font-size: 14px;
    }

    /* -----------------------------
       DOWNLOAD BUTTON
    ----------------------------- */
    div[data-testid="stDownloadButton"] button {
        background-color: #22c55e !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: bold !important;
        background-image: none !important;
    }

    div[data-testid="stDownloadButton"] button:hover {
        background-color: #16a34a !important;
        color: white !important;
    }

    div[data-testid="stDownloadButton"] button:focus,
    div[data-testid="stDownloadButton"] button:active {
        background-color: #16a34a !important;
        color: white !important;
        outline: none !important;
        box-shadow: none !important;
    }

    div[data-testid="stDownloadButton"] button span {
        color: white !important;
    }

</style>
""", unsafe_allow_html=True)

st.title("DLP LLM Proxy Dashboard")
st.markdown(
    "<p class='small-note'>Live view of mitmproxy traffic inspected by the DLP engine.</p>",
    unsafe_allow_html=True
)

# -----------------------------
# Refresh + File Check
# -----------------------------
refresh_rate = st.sidebar.slider(
    "Auto Refresh (seconds)",
    2,
    30,
    5
)

if not os.path.exists(LOG_FILE):
    st.warning("No log file found yet. Run mitmproxy traffic first.")
    st.stop()

# -----------------------------
# Load Data
# -----------------------------
@st.cache_data(ttl=2)
def load_data():
    df = pd.read_csv(LOG_FILE)

    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
        df["timestamp"] = df["timestamp"].dt.tz_convert("America/New_York")

    return df

@st.cache_data
def convert_df_to_excel(df_in):
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    export_df = df_in.copy()

    # Add export-friendly content columns so the Excel file includes readable content.
    if "preview" not in export_df.columns:
        export_df["preview"] = ""
    if "full_content" not in export_df.columns:
        export_df["full_content"] = ""

    def _preview_for_export(row):
        for col in ["match_excerpt", "redacted_excerpt", "preview", "extracted_text", "original_excerpt"]:
            if col in row.index and pd.notna(row[col]) and str(row[col]).strip():
                val = str(row[col])
                return val[:5000]
        return ""

    def _full_for_export(row):
        for col in ["extracted_text", "original_excerpt", "match_excerpt", "redacted_excerpt"]:
            if col in row.index and pd.notna(row[col]) and str(row[col]).strip():
                return str(row[col])[:20000]
        return ""

    export_df["preview"] = export_df.apply(_preview_for_export, axis=1)
    export_df["full_content"] = export_df.apply(_full_for_export, axis=1)

    for col in export_df.columns:
        if pd.api.types.is_datetime64tz_dtype(export_df[col]):
            export_df[col] = (
                export_df[col]
                .dt.tz_localize(None)
                .dt.strftime("%Y-%m-%d %H:%M:%S")
            )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="DLP Events")

        workbook = writer.book
        worksheet = writer.sheets["DLP Events"]

        fills = {
            "ALLOW": PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),
            "COACH": PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"),
            "BLOCK": PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
            "QUARANTINE": PatternFill(fill_type="solid", start_color="E4C1F9", end_color="E4C1F9"),
        }

        header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        header_font = Font(bold=True)

        # Style header row.
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top")

        # Freeze top row and enable filters.
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        action_col_idx = None
        for idx, col_name in enumerate(export_df.columns, start=1):
            if col_name == "action":
                action_col_idx = idx
                break

        # Color each row based on action.
        if action_col_idx is not None:
            for row_idx in range(2, len(export_df) + 2):
                action_val = worksheet.cell(row=row_idx, column=action_col_idx).value
                fill = fills.get(str(action_val).upper())
                if fill:
                    for col_idx in range(1, len(export_df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill

        # Wrap long text columns and set widths.
        long_text_cols = {"preview", "full_content", "match_excerpt", "extracted_text", "original_excerpt", "reasons", "reason"}
        for idx, col_name in enumerate(export_df.columns, start=1):
            letter = get_column_letter(idx)
            if col_name in long_text_cols:
                worksheet.column_dimensions[letter].width = 50
                for row_idx in range(2, len(export_df) + 2):
                    worksheet.cell(row=row_idx, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
            else:
                max_len = max([len(str(col_name))] + [len(str(v)) for v in export_df[col_name].head(50).fillna("")])
                worksheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 28)

    return output.getvalue()

df = load_data()

if df.empty:
    st.info("Log file exists, but no events have been recorded yet.")
    time.sleep(refresh_rate)
    st.rerun()

# -----------------------------
# Helpers
# -----------------------------
ACTION_COLORS = {
    "ALLOW": "#22c55e",
    "COACH": "#f59e0b",
    "BLOCK": "#ef4444",
    "QUARANTINE": "#a855f7"
}

LABEL_COLORS = {
    "PII": "#ef4444",
    "CONFIDENTIAL": "#f59e0b",
    "BENIGN": "#22c55e",
    "UNKNOWN": "#38bdf8"
}

def style_action(val):
    val = str(val).upper()
    if val == "ALLOW":
        return "background-color: #14532d; color: white;"
    if val == "COACH":
        return "background-color: #78350f; color: white;"
    if val == "BLOCK":
        return "background-color: #7f1d1d; color: white;"
    if val == "QUARANTINE":
        return "background-color: #581c87; color: white;"
    return ""

def safe_col_exists(df_in, col_name):
    return col_name in df_in.columns

def get_preview_text(row):
    # ✅ prefer new match_excerpt
    if "match_excerpt" in row.index and pd.notna(row["match_excerpt"]) and str(row["match_excerpt"]).strip():
        return str(row["match_excerpt"])
    # existing fallbacks
    if "redacted_excerpt" in row.index and pd.notna(row["redacted_excerpt"]):
        return str(row["redacted_excerpt"])
    if "preview" in row.index and pd.notna(row["preview"]):
        return str(row["preview"])
    if "extracted_text" in row.index and pd.notna(row["extracted_text"]):
        txt = str(row["extracted_text"])
        return txt[:120] + ("..." if len(txt) > 120 else "")
    return ""

def get_full_text(row):
    if "extracted_text" in row.index and pd.notna(row["extracted_text"]):
        return str(row["extracted_text"])
    if "original_excerpt" in row.index and pd.notna(row["original_excerpt"]):
        return str(row["original_excerpt"])
    return ""

# -----------------------------
# Sidebar Filters
# -----------------------------
st.sidebar.header("Filters")

action_options = sorted(df["action"].dropna().unique().tolist()) if safe_col_exists(df, "action") else []
host_options = sorted(df["host"].dropna().unique().tolist()) if safe_col_exists(df, "host") else []
label_options = sorted(df["label"].dropna().unique().tolist()) if safe_col_exists(df, "label") else []

action_filter = st.sidebar.multiselect(
    "Action",
    action_options,
    default=action_options
)

host_filter = st.sidebar.multiselect(
    "Host",
    host_options,
    default=host_options
)

label_filter = st.sidebar.multiselect(
    "Label",
    label_options,
    default=label_options
)

filtered_df = df.copy()

if safe_col_exists(filtered_df, "action"):
    filtered_df = filtered_df[filtered_df["action"].isin(action_filter)]

if safe_col_exists(filtered_df, "host"):
    filtered_df = filtered_df[filtered_df["host"].isin(host_filter)]

if safe_col_exists(filtered_df, "label"):
    filtered_df = filtered_df[filtered_df["label"].isin(label_filter)]

# -----------------------------
# Timestamp Filter
# -----------------------------
if safe_col_exists(filtered_df, "timestamp") and not filtered_df.empty:
    min_ts = filtered_df["timestamp"].min()
    max_ts = filtered_df["timestamp"].max()

    if pd.notna(min_ts) and pd.notna(max_ts):
        time_range = st.sidebar.slider(
            "Timestamp Range",
            min_value=min_ts.to_pydatetime(),
            max_value=max_ts.to_pydatetime(),
            value=(min_ts.to_pydatetime(), max_ts.to_pydatetime()),
            format="MM/DD/YY - HH:mm"
        )

        filtered_df = filtered_df[
            (filtered_df["timestamp"] >= pd.Timestamp(time_range[0])) &
            (filtered_df["timestamp"] <= pd.Timestamp(time_range[1]))
        ]

# -----------------------------
# Tabs
# -----------------------------
overview_tab, quarantine_tab = st.tabs(["Overview", "Quarantine Review"])

# =========================================================
# OVERVIEW TAB
# =========================================================
with overview_tab:
    total_events = len(filtered_df)
    blocked_count = len(filtered_df[filtered_df["action"] == "BLOCK"]) if safe_col_exists(filtered_df, "action") else 0
    coached_count = len(filtered_df[filtered_df["action"] == "COACH"]) if safe_col_exists(filtered_df, "action") else 0
    allowed_count = len(filtered_df[filtered_df["action"] == "ALLOW"]) if safe_col_exists(filtered_df, "action") else 0
    quarantined_count = len(filtered_df[filtered_df["action"] == "QUARANTINE"]) if safe_col_exists(filtered_df, "action") else 0

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total Events", total_events)
    col2.metric("Blocked", blocked_count)
    col3.metric("Coached", coached_count)
    col4.metric("Allowed", allowed_count)
    col5.metric("Quarantined", quarantined_count)

    st.divider()

    excel_data = convert_df_to_excel(filtered_df)

    st.download_button(
        label="Export to Excel",
        key="overview_export",
        data=excel_data,
        file_name="dlp_dashboard_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.subheader("Action Distribution")

        if safe_col_exists(filtered_df, "action") and not filtered_df.empty:
            action_counts = filtered_df["action"].value_counts()
            bar_colors = [ACTION_COLORS.get(a, "#38bdf8") for a in action_counts.index]

            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor("#0f172a")
            ax.set_facecolor("#082f49")

            action_counts.plot(
                kind="bar",
                ax=ax,
                color=bar_colors,
                edgecolor="white"
            )

            ax.set_title("Policy Actions", color="white", fontsize=14, pad=12)
            ax.set_xlabel("Action", color="white")
            ax.set_ylabel("Count", color="white")
            ax.tick_params(colors="white")
            ax.spines["bottom"].set_color("white")
            ax.spines["left"].set_color("white")
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.grid(axis="y", linestyle="--", alpha=0.3)

            st.pyplot(fig)
        else:
            st.info("No action data available.")

    with chart_col2:
        st.subheader("Label Distribution")

        if safe_col_exists(filtered_df, "label") and not filtered_df.empty:
            label_counts = filtered_df["label"].value_counts()
            pie_colors = [LABEL_COLORS.get(lbl, "#67e8f9") for lbl in label_counts.index]

            fig2, ax2 = plt.subplots(figsize=(5, 5))
            fig2.patch.set_facecolor("#0f172a")
            ax2.set_facecolor("#082f49")

            wedges, texts, autotexts = ax2.pie(
                label_counts,
                labels=label_counts.index,
                autopct="%1.1f%%",
                colors=pie_colors,
                textprops={"color": "white"}
            )

            ax2.set_title("Detected Content Types", color="white", fontsize=14, pad=12)

            for autotext in autotexts:
                autotext.set_color("white")

            st.pyplot(fig2)
        else:
            st.info("No label data available.")

    st.subheader("Traffic Trend")

    if safe_col_exists(filtered_df, "timestamp") and not filtered_df.empty:
        trend_df = filtered_df.copy()
        trend_df["minute"] = trend_df["timestamp"].dt.floor("min")
        trend_counts = trend_df.groupby("minute").size()

        fig3, ax3 = plt.subplots(figsize=(10, 4))
        fig3.patch.set_facecolor("#0f172a")
        ax3.set_facecolor("#082f49")

        ax3.plot(trend_counts.index, trend_counts.values)
        ax3.set_title("Event Volume by Minute", color="white", fontsize=14, pad=12)
        ax3.set_xlabel("Time", color="white")
        ax3.set_ylabel("Events", color="white")
        ax3.tick_params(colors="white")
        ax3.spines["bottom"].set_color("white")
        ax3.spines["left"].set_color("white")
        ax3.spines["top"].set_visible(False)
        ax3.spines["right"].set_visible(False)
        ax3.grid(axis="y", linestyle="--", alpha=0.3)

        st.pyplot(fig3)
    else:
        st.info("No timestamp data available for trend graph.")

    st.divider()

    left, right = st.columns([2, 1])

    with left:
        st.subheader("Recent Events")

        display_df = filtered_df.copy()

        # ✅ prefer match_excerpt > redacted_excerpt > extracted_text
        if safe_col_exists(display_df, "match_excerpt"):
            display_df["preview"] = display_df["match_excerpt"].astype(str)
        elif safe_col_exists(display_df, "redacted_excerpt"):
            display_df["preview"] = display_df["redacted_excerpt"].astype(str)
        elif safe_col_exists(display_df, "extracted_text"):
            display_df["preview"] = display_df["extracted_text"].astype(str).str[:100] + "..."
        else:
            display_df["preview"] = ""

        display_cols = []
        for col in [
            "timestamp",
            "host",
            "destination",
            "label",
            "action",
            "risk_score",
            # ✅ new columns
            "match_type",
            "preview"
        ]:
            if col in display_df.columns:
                display_cols.append(col)

        if display_cols:
            recent_df = display_df.sort_values("timestamp", ascending=False) if safe_col_exists(display_df, "timestamp") else display_df.copy()
            recent_df = recent_df[display_cols].head(25)

            try:
                styled_recent = recent_df.style.map(style_action, subset=["action"]) if "action" in recent_df.columns else recent_df.style
                st.dataframe(styled_recent, use_container_width=True, height=500)
            except Exception:
                st.dataframe(recent_df, use_container_width=True, height=500)
        else:
            st.info("No columns available to display.")

    with right:
        st.subheader("Event Preview")

        preview_df = filtered_df.sort_values("timestamp", ascending=False).reset_index(drop=True) if safe_col_exists(filtered_df, "timestamp") else filtered_df.reset_index(drop=True)

        if len(preview_df) > 0:
            selected_idx = st.number_input(
                "Select row for preview",
                min_value=0,
                max_value=len(preview_df) - 1,
                value=0,
                step=1
            )

            row = preview_df.iloc[int(selected_idx)]

            st.write("**Destination:**", row.get("destination", ""))
            st.write("**Host:**", row.get("host", ""))
            st.write("**Path:**", row.get("path", ""))
            st.write("**Label:**", row.get("label", ""))
            st.write("**Action:**", row.get("action", ""))
            st.write("**Risk Score:**", row.get("risk_score", ""))

            # ✅ new fields
            st.write("**Match Type:**", row.get("match_type", ""))
            st.write("**Match Excerpt:**")
            match_ex = row.get("match_excerpt", "")
            if pd.notna(match_ex) and str(match_ex).strip():
                st.code(str(match_ex), language="text")
            else:
                st.caption("No match excerpt (no pattern hit or field not present).")

            st.write("**Confidence:**", row.get("confidence", ""))
            st.write("**Decision Source:**", row.get("decision_source", ""))

            reveal = st.checkbox("👁 Reveal sensitive details", value=True)

            st.write("**Preview:**")
            if reveal:
                full_text = get_full_text(row)
                if full_text:
                    st.code(full_text, language="text")
                else:
                    st.info("No unredacted content stored for this event.")
            else:
                preview_text = get_preview_text(row)
                if preview_text:
                    st.code(preview_text, language="text")
                else:
                    st.info("No preview available.")

            if str(row.get("action", "")).upper() == "QUARANTINE":
                st.write("### Quarantine Evidence")

                if "payload_sha256" in row.index and pd.notna(row["payload_sha256"]):
                    st.write("**SHA-256:**", row.get("payload_sha256", ""))

                if "reasons" in row.index and pd.notna(row["reasons"]):
                    st.write("**Reasons:**", row.get("reasons"))
                elif "reason" in row.index and pd.notna(row["reason"]):
                    st.write("**Reason:**", row.get("reason"))

                file_path = row.get("stored_file_path", "")
                file_name = row.get("original_filename", "quarantined_file.bin")

                if file_path and os.path.exists(file_path):
                    with open(file_path, "rb") as f:
                        st.download_button(
                            label="Download Attachment",
                            data=f,
                            file_name=file_name,
                            mime="application/octet-stream"
                        )
                else:
                    st.info("No stored attachment available for download.")
        else:
            st.info("No events available for preview.")

# =========================================================
# QUARANTINE REVIEW TAB
# =========================================================
with quarantine_tab:
    st.subheader("Quarantine Review Queue")

    if safe_col_exists(filtered_df, "action"):
        quarantined_df = filtered_df[filtered_df["action"] == "QUARANTINE"].copy()
    else:
        quarantined_df = pd.DataFrame()

    if not quarantined_df.empty:
        quarantine_excel_data = convert_df_to_excel(quarantined_df)

        st.download_button(
            label="Export Quarantine to Excel",
            key="quarantine_export_excel",
            data=quarantine_excel_data,
            file_name="dlp_quarantine_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if quarantined_df.empty:
        st.info("No quarantined events found in the current log.")
    else:
        q1, q2, q3 = st.columns(3)
        q1.metric("Quarantined Events", len(quarantined_df))

        left, right = st.columns([2, 1])

        with left:
            review_cols = []
            for col in [
                "timestamp",
                "host",
                "destination",
                "path",
                "label",
                "action",
                "risk_score",
                # ✅ new columns
                "match_type",
                "match_excerpt",
                "payload_sha256"
            ]:
                if col in quarantined_df.columns:
                    review_cols.append(col)

            st.dataframe(
                quarantined_df.sort_values("timestamp", ascending=False)[review_cols],
                use_container_width=True,
                height=500
            )

        with right:
            st.write("### Quarantine Detail")

            q_df = quarantined_df.sort_values("timestamp", ascending=False).reset_index(drop=True)
            selected_q_idx = st.number_input(
                "Select quarantined row",
                min_value=0,
                max_value=len(q_df) - 1,
                value=0,
                step=1,
                key="q_row"
            )

            row = q_df.iloc[int(selected_q_idx)]

            st.write("**Timestamp:**", row.get("timestamp", ""))
            st.write("**Host:**", row.get("host", ""))
            st.write("**Destination:**", row.get("destination", ""))
            st.write("**Path:**", row.get("path", ""))
            st.write("**Label:**", row.get("label", ""))
            st.write("**Action:**", row.get("action", ""))
            st.write("**Risk Score:**", row.get("risk_score", ""))

            # ✅ new fields
            st.write("**Match Type:**", row.get("match_type", ""))
            st.write("**Match Excerpt:**")
            match_ex = row.get("match_excerpt", "")
            if pd.notna(match_ex) and str(match_ex).strip():
                st.code(str(match_ex), language="text")
            else:
                st.caption("No match excerpt (no pattern hit or field not present).")

            st.write("**Confidence:**", row.get("confidence", ""))
            st.write("**Decision Source:**", row.get("decision_source", ""))

            if "payload_sha256" in row.index and pd.notna(row["payload_sha256"]):
                st.write("**SHA-256:**", row.get("payload_sha256", ""))

            reveal_q = st.checkbox("👁 Reveal sensitive details", value=True, key="q_reveal")

            st.write("**Excerpt:**")
            if reveal_q:
                full_text = get_full_text(row)
                if full_text:
                    st.code(full_text, language="text")
                else:
                    st.info("No unredacted content stored for this event.")
            else:
                preview_text = get_preview_text(row)
                if preview_text:
                    st.code(preview_text, language="text")
                else:
                    st.info("No preview available.")

            if "reasons" in row.index and pd.notna(row["reasons"]):
                st.write("**Reasons:**", row.get("reasons"))
            elif "reason" in row.index and pd.notna(row["reason"]):
                st.write("**Reason:**", row.get("reason"))

            file_path = row.get("stored_file_path", "")
            file_name = row.get("original_filename", "quarantined_file.bin")

            if file_path and os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    st.download_button(
                        label="Download Attachment",
                        data=f,
                        file_name=file_name,
                        mime="application/octet-stream",
                        key="q_download"
                    )
            else:
                st.info("No stored attachment available for download.")

            st.write("### Full Event JSON")
            st.json(row.to_dict())

# -----------------------------
# Auto Refresh
# -----------------------------
time.sleep(refresh_rate)
st.rerun()
